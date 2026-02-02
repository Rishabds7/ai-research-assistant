"""
Export utilities: Excel export for methodologies and research gaps.
Uses openpyxl.
"""

from pathlib import Path
from typing import Any, Optional, Union

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def _flatten_methodology(m: dict[str, Any], paper_id: str = "") -> dict[str, Any]:
    """Flatten one methodology dict for table row."""
    row: dict[str, Any] = {"Paper": paper_id}
    datasets = m.get("datasets") or []
    if isinstance(datasets, list) and datasets:
        ds = datasets[0]
        if isinstance(ds, dict):
            row["Dataset"] = ds.get("name") or ""
            row["Dataset Size"] = ds.get("size") or ""
            row["Dataset Source"] = ds.get("source") or ""
        else:
            row["Dataset"] = str(ds)
    else:
        row["Dataset"] = ""
        row["Dataset Size"] = ""
        row["Dataset Source"] = ""
    model = m.get("model") or {}
    if isinstance(model, dict):
        row["Model"] = model.get("name") or ""
        row["Model Architecture"] = model.get("architecture") or ""
    else:
        row["Model"] = str(model)
    metrics = m.get("metrics") or []
    row["Metrics"] = ", ".join(str(x) for x in metrics) if metrics else ""
    results = m.get("results") or {}
    row["Results"] = str(results)[:500] if results else ""
    row["Contribution"] = (m.get("contribution") or "")[:500]
    return row


def export_to_excel(
    methodologies: list[dict[str, Any]],
    gaps: Optional[dict[str, Any]],
    output_path: Union[Path, str],
) -> str:
    """
    Create an Excel file with two sheets: Methodologies and Research Gaps.

    Args:
        methodologies: List of methodology dicts (each may have paper_id in structure
                       or pass list of (paper_id, methodology) - we assume list of
                       dicts; if dict has 'paper_id' use it, else use index).
        gaps: Dict with methodological_gaps, dataset_limitations, evaluation_gaps,
              novel_directions (lists of dicts).
        output_path: Path for the output .xlsx file.

    Returns:
        Absolute path to the created file as string.
    """
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Methodologies"

    # Sheet 1: Methodologies
    if methodologies:
        # Assume each item is dict; paper_id might be in a wrapper - we accept
        # list of dicts where we use index as paper id if no paper_id
        headers = [
            "Paper",
            "Dataset",
            "Dataset Size",
            "Dataset Source",
            "Model",
            "Model Architecture",
            "Metrics",
            "Results",
            "Contribution",
        ]
        for col, h in enumerate(headers, 1):
            ws1.cell(row=1, column=col, value=h)
        for row_idx, m in enumerate(methodologies, 2):
            paper_id = m.get("paper_id", f"Paper_{row_idx - 1}")
            flat = _flatten_methodology(m, paper_id)
            for col, key in enumerate(headers, 1):
                val = flat.get(key, "")
                ws1.cell(row=row_idx, column=col, value=val)
    else:
        ws1.cell(row=1, column=1, value="No methodologies")

    # Sheet 2: Research Gaps
    ws2 = wb.create_sheet("Research Gaps")
    if gaps:
        row = 1
        for section in (
            "methodological_gaps",
            "dataset_limitations",
            "evaluation_gaps",
            "novel_directions",
        ):
            items = gaps.get(section) or []
            if not isinstance(items, list):
                continue
            ws2.cell(row=row, column=1, value=section.replace("_", " ").title())
            row += 1
            for item in items:
                if isinstance(item, dict):
                    desc = item.get("description") or item.get("value") or str(item)
                else:
                    desc = str(item)
                ws2.cell(row=row, column=1, value=desc)
                row += 1
            row += 1
    else:
        ws2.cell(row=1, column=1, value="No gaps data")

    wb.save(path)
    return str(path.resolve())
